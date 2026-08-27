import openpyxl
from openpyxl.styles import PatternFill, Font

bold_font = Font(bold=True)
redFill = PatternFill(start_color='FF7F7F', end_color='FF7F7F', fill_type='solid')
blueFill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

def addToSheet(file_name, data, sheet_name):
    workbook = openpyxl.load_workbook(file_name)
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    sheet = workbook.create_sheet(sheet_name)
    for row_index, row in enumerate(data):
        for col_index, cell_value in enumerate(row):
            cell = sheet.cell(row=row_index + 1, column=col_index + 1, value=cell_value)
            if row_index == 0: cell.font = bold_font
            if cell_value == "Loss" or cell_value == "Not Available": cell.fill = redFill
            elif cell_value == 'Gain': cell.fill = blueFill
    wb = workbook.active
    for col in wb.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 2
        wb.column_dimensions[column].width = adjusted_width
        sheet.auto_filter.ref = sheet.dimensions
    workbook.save(file_name)